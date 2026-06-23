"""
Módulo do bot de criação de pedidos no Bling.
Encapsula toda a lógica em uma classe para uso pela interface web.
"""

import io
import os
import re
import shutil
import time
import unicodedata
from datetime import datetime
from typing import Callable

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait

# ================== CONSTANTES ==================
URL_LOGIN = "https://www.bling.com.br/b/login.php"
URL_PEDIDOS = "https://www.bling.com.br/b/vendas.php#add"
TIMEOUT = 15
NOME_ABA = "Montagem de Pedidos - Envio"
AUTOCOMPLETE_XPATH = (
    "//ul[contains(@class,'ui-autocomplete')]"
    "//li[not(contains(@class,'ui-autocomplete-empty'))]"
)
_ERRO_NUMERO_KEYWORDS = [
    "numero ja", "número já", "numero existe", "número existe",
    "ja cadastrado", "já cadastrado", "duplicado", "ja foi utilizado",
    "já foi utilizado", "numero informado", "número informado",
]
_FILL_AMARELO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_FILL_VERMELHO = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")


def _strip_accents(text: str) -> str:
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _normalize(text: str) -> str:
    return re.sub(r"\s+", "", _strip_accents(text)).upper() if text else ""


class BlingBot:
    """Bot de criação de pedidos no Bling com callbacks de log."""

    def __init__(self, caminho_planilha: str, resultado_dir: str, prints_dir: str,
                 log: Callable[[str], None] | None = None,
                 email: str | None = None, senha: str | None = None,
                 cliente: str | None = None, loja: str | None = None,
                 frete: str | None = None, data_prevista: str | None = None):
        load_dotenv()
        self.email = email or os.getenv("BLING_EMAIL")
        self.senha = senha or os.getenv("BLING_SENHA")
        if not self.email or not self.senha:
            raise ValueError("Informe e-mail e senha do Bling (no formulário ou no arquivo .env).")

        self.cliente = (cliente or "ENVIO FULL ML").strip()
        self.loja = (loja or "Loja Full").strip()
        self.frete = (frete or "9 - Sem Ocorrência de Transporte").strip()
        self.data_prevista = self._normalizar_data(data_prevista)

        self.caminho_planilha = caminho_planilha
        self.resultado_dir = resultado_dir
        self.prints_dir = prints_dir
        self._log = log or print
        self._cancelado = False

        self.driver = None
        self.wait = None
        self.wait_short = None

        self.pedidos: list[list[dict]] = []
        self.ok_pedidos: list[int] = []
        self.err_pedidos: list[int] = []
        self._alteracoes: list[dict] = []
        self.resultado_filename: str | None = None
        self.resultado_bytes: bytes | None = None
        self.caminho_resultado: str | None = None  # mantido por compatibilidade (= filename)
        self._header_row: int = 1  # linha (1-based) do cabeçalho no Excel — detectada

    # ---------- utilidades ----------
    def log(self, msg: str):
        self._log(msg)

    @staticmethod
    def _normalizar_data(valor: str | None) -> str | None:
        """Converte data para o formato dd/mm/yyyy esperado pelo Bling.
        Aceita 'yyyy-mm-dd' (input HTML) ou já no formato BR. Retorna None se vazio/invalido."""
        if not valor:
            return None
        s = str(valor).strip()
        if not s:
            return None
        # Formato HTML <input type="date">: yyyy-mm-dd
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
        if m:
            return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
        # Já em dd/mm/yyyy ou dd-mm-yyyy
        m = re.match(r"^(\d{2})[/-](\d{2})[/-](\d{4})$", s)
        if m:
            return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
        return s  # último recurso — manda como veio

    def cancelar(self):
        """Cancela imediatamente: marca flag e encerra o navegador para
        interromper qualquer chamada Selenium em andamento."""
        self._cancelado = True
        self.log("🛑 Cancelando execução...")
        drv = self.driver
        if drv is not None:
            try:
                drv.quit()
            except Exception:
                pass
            self.driver = None

    def _click_js(self, el):
        self.driver.execute_script("arguments[0].click();", el)

    def _clicar_botao_ok(self, timeout=6):
        try:
            btn = WebDriverWait(self.driver, timeout).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//div[contains(@class,'ui-dialog')]//button[normalize-space()='Ok']",
                ))
            )
            self._click_js(btn)
            return True
        except Exception:
            pass
        try:
            btn = WebDriverWait(self.driver, 1).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//button[contains(normalize-space(.),'Ok') or contains(normalize-space(.),'OK')]",
                ))
            )
            self._click_js(btn)
            return True
        except Exception:
            return False

    def _capturar_numero_pedido(self) -> str | None:
        try:
            h1 = self.driver.find_element(By.ID, "saleOrderHeader")
            match = re.search(r"-\s*(\d+)", h1.text.strip())
            if match:
                return match.group(1)
        except Exception:
            pass
        return None

    def _first_clickable(self, selectors, timeout=10):
        for by, sel in selectors:
            try:
                return WebDriverWait(self.driver, timeout).until(
                    EC.element_to_be_clickable((by, sel))
                )
            except Exception:
                continue
        raise RuntimeError("Elemento não encontrado na página.")

    # ---------- planilha ----------
    def _detectar_linha_cabecalho(self) -> int:
        """Procura nas primeiras linhas a célula 'SKU PAI' e devolve a linha
        (1-based) onde o cabeçalho está. Robusto a banners/títulos no topo."""
        wb = load_workbook(self.caminho_planilha, read_only=True, data_only=True)
        try:
            ws = wb[NOME_ABA]
            limite = min(30, ws.max_row or 30)
            for row_idx in range(1, limite + 1):
                for col_idx in range(1, (ws.max_column or 0) + 1):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val and _normalize(str(val)) == "SKUPAI":
                        return row_idx
        finally:
            wb.close()
        raise ValueError(
            "Não encontrei a coluna 'SKU PAI' nas primeiras 30 linhas da aba "
            f"'{NOME_ABA}'."
        )

    def carregar_planilha(self):
        self._header_row = self._detectar_linha_cabecalho()
        self.log(f"📑 Cabeçalho detectado na linha {self._header_row}.")

        df = pd.read_excel(
            self.caminho_planilha,
            sheet_name=NOME_ABA,
            header=self._header_row - 1,
        )
        # Normaliza nomes (remove espaços extras)
        df.columns = [str(c).strip() for c in df.columns]

        if not all(col in df.columns for col in ["SKU PAI", "QNT PLANEJADA"]):
            raise ValueError("A planilha deve conter as colunas 'SKU PAI' e 'QNT PLANEJADA'.")

        self.pedidos = []
        pedido_atual: list[dict] = []

        for idx, row in df.iterrows():
            sku = str(row["SKU PAI"]).strip() if pd.notna(row["SKU PAI"]) else ""
            qty = row["QNT PLANEJADA"]
            # Linha do Excel = índice pandas + linha-do-cabeçalho + 1
            excel_row = int(idx) + self._header_row + 1
            if pd.isna(qty) or sku == "":
                if pedido_atual:
                    self.pedidos.append(pedido_atual)
                    pedido_atual = []
            else:
                pedido_atual.append({"SKU PAI": sku, "QNT PLANEJADA": int(qty), "excel_row": excel_row})

        if pedido_atual:
            self.pedidos.append(pedido_atual)

        # Define nome do arquivo de resultado (gerado em memória, sem salvar no disco)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_base = os.path.splitext(os.path.basename(self.caminho_planilha))[0]
        self.resultado_filename = f"{nome_base}_resultado_{timestamp}.xlsx"
        self.caminho_resultado = self.resultado_filename  # compat

        self.log(f"✅ {len(self.pedidos)} pedido(s) carregado(s) da planilha.")
        for i, p in enumerate(self.pedidos, 1):
            skus = ", ".join(x["SKU PAI"] for x in p)
            self.log(f"   Pedido {i}: {len(p)} produto(s) → {skus}")

    def gravar_resultado_final(self):
        if not self._alteracoes:
            self.log("ℹ️ Nenhuma alteração para gravar.")
            return

        wb = load_workbook(self.caminho_planilha)
        ws = wb[NOME_ABA]

        col_num = None
        col_qtd = None
        for col in range(1, ws.max_column + 1):
            h = str(ws.cell(row=self._header_row, column=col).value or "").upper()
            if col_num is None and "PEDIDO" in h and "VENDA" in h and "QNT" not in h:
                col_num = col
            if col_qtd is None and "QNT" in h and "PEDIDO" in h and "VENDA" in h:
                col_qtd = col

        col_num = col_num or 2
        col_qtd = col_qtd or 8

        for alt in self._alteracoes:
            linha = alt["linha"]
            if alt.get("numero_pedido"):
                ws.cell(row=linha, column=col_num, value=alt["numero_pedido"])
            if alt.get("qtd_real") is not None:
                cell = ws.cell(row=linha, column=col_qtd, value=alt["qtd_real"])
                obs = alt.get("obs", "")
                if "SEM ESTOQUE" in obs:
                    cell.fill = _FILL_VERMELHO
                elif obs:
                    cell.fill = _FILL_AMARELO

        # Serializa em memória (BytesIO) — não grava no disco
        buf = io.BytesIO()
        wb.save(buf)
        self.resultado_bytes = buf.getvalue()
        self.log(f"📝 Planilha pronta para download: {self.resultado_filename}")

    # ---------- login ----------
    def _switch_into_login_iframe(self):
        xpath = (
            "//input[@type='email' or @name='email'"
            " or contains(@placeholder,'E-mail') or contains(@placeholder,'usuário')]"
        )
        try:
            for fr in self.driver.find_elements(By.TAG_NAME, "iframe"):
                self.driver.switch_to.default_content()
                try:
                    self.driver.switch_to.frame(fr)
                    self.driver.find_element(By.XPATH, xpath)
                    self.log("🧩 Login dentro de iframe detectado.")
                    return True
                except Exception:
                    continue
            self.driver.switch_to.default_content()
        except Exception:
            self.driver.switch_to.default_content()
        return False

    def _preencher_login(self):
        email_input = self._first_clickable([
            (By.XPATH, "//input[@type='email' or @name='email' or contains(@placeholder,'E-mail') or contains(@placeholder,'usuário')]"),
            (By.CSS_SELECTOR, "input[type='email'], input[name='email']"),
        ])
        email_input.clear()
        email_input.send_keys(self.email)

        senha_input = self._first_clickable([
            (By.XPATH, "//input[@type='password' or @name='senha']"),
            (By.CSS_SELECTOR, "input[type='password']"),
        ])
        senha_input.clear()
        senha_input.send_keys(self.senha)

        botao = self._first_clickable([
            (By.XPATH, "//button[@type='submit' or contains(.,'Entrar') or contains(.,'Acessar')]"),
            (By.CSS_SELECTOR, "form button[type='submit']"),
        ])
        self._click_js(botao)

    def fazer_login(self):
        self.log("➡️ Fazendo login no Bling…")
        self.driver.get(URL_LOGIN)
        self.wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(2)
        self._switch_into_login_iframe()

        try:
            self._preencher_login()
        except Exception:
            self.driver.switch_to.default_content()
            self._preencher_login()

        try:
            self.wait.until(EC.url_contains("inicio"))
        except Exception:
            self.driver.get("https://www.bling.com.br/inicio#/")

        self.driver.switch_to.default_content()
        self.log("✅ Login realizado com sucesso!")

    # ---------- SKU autocomplete ----------
    def _selecionar_sku(self, campo, produto_nome: str) -> bool:
        campo.clear()
        campo.send_keys(produto_nome)
        time.sleep(3)  # dá tempo do Bling consultar e popular o autocomplete

        try:
            self.wait_short.until(EC.visibility_of_element_located((By.XPATH, AUTOCOMPLETE_XPATH)))
        except Exception:
            campo.send_keys(Keys.ENTER)
            return False

        itens = self.driver.find_elements(By.XPATH, AUTOCOMPLETE_XPATH)
        sku_busca = _normalize(produto_nome)
        candidatos = []

        for it in itens:
            if not it.is_displayed():
                continue

            sku_val = ""
            descricao = ""
            span_cod = None

            try:
                try:
                    span_cod = it.find_element(
                        By.XPATH,
                        ".//span[contains(@class,'ui-autocomplete-custom-item')"
                        " and (contains(.,'Cód') or contains(.,'Cod'))]",
                    )
                except Exception:
                    for sp in it.find_elements(By.TAG_NAME, "span"):
                        if sp.text and "cod" in _strip_accents(sp.text).lower():
                            span_cod = sp
                            break

                if span_cod is not None:
                    txt = span_cod.text.strip()
                    idx = max(txt.rfind(":"), txt.rfind("："))
                    sku_val = (
                        txt[idx + 1:].strip()
                        if idx != -1
                        else re.sub(r"(?i)cod\.?\s*[:\uFF1A]?\s*", "", txt).strip()
                    )
                else:
                    spans = it.find_elements(By.TAG_NAME, "span")
                    if spans:
                        descricao = spans[0].text.strip()
            except Exception:
                pass

            sku_norm = _normalize(sku_val)
            desc_norm = _normalize(descricao)
            candidatos.append(sku_norm or desc_norm or "(sem)")

            if sku_norm == sku_busca or desc_norm == sku_busca:
                try:
                    target = span_cod if span_cod is not None else it
                    try:
                        self._click_js(target)
                    except Exception:
                        try:
                            self._click_js(target.find_element(By.XPATH, "./ancestor::a[1]"))
                        except Exception:
                            self._click_js(it)
                    return True
                except Exception:
                    try:
                        self._click_js(it)
                        return True
                    except Exception:
                        pass

        self.log(f"⚠️ SKU '{produto_nome}' não encontrado. Candidatos: {', '.join(candidatos[:10])}")
        campo.send_keys(Keys.ENTER)
        return False

    # ---------- validação de estoque ----------
    def _achar_campo_estoque(self, idx_item: int):
        """Localiza o campo oculto de estoque tentando vários padrões de ID.
        O Bling pode usar índice 0-based, 1-based, ou sem sufixo na 1ª linha."""
        candidatos = [
            f"h_estoque_atual_{idx_item}",
            f"h_estoque_atual_{idx_item + 1}",
        ]
        if idx_item == 0:
            candidatos.append("h_estoque_atual")
        for cid in candidatos:
            try:
                return self.driver.find_element(By.ID, cid)
            except Exception:
                continue
        # Fallback: pega o último input cujo id começa com h_estoque_atual (mais recente = nova linha)
        try:
            els = self.driver.find_elements(
                By.XPATH, "//input[starts-with(@id,'h_estoque_atual')]"
            )
            if els:
                return els[-1]
        except Exception:
            pass
        return None

    def _ler_estoque_estavel(self, el, timeout: float = 5.0) -> int | None:
        """Aguarda até que o valor do campo de estoque pare de mudar e o retorna.
        Retorna None se não conseguir ler de forma confiável."""
        deadline = time.time() + timeout
        anterior = None
        ultimo_valido = None
        estavel_desde = None

        while time.time() < deadline:
            try:
                v = (el.get_attribute("value") or "").strip()
            except Exception:
                return None

            if v != "":
                if v == anterior:
                    if estavel_desde is None:
                        estavel_desde = time.time()
                    elif time.time() - estavel_desde >= 0.8:
                        ultimo_valido = v
                        break
                else:
                    estavel_desde = None
                anterior = v
            time.sleep(0.25)

        if ultimo_valido is None:
            ultimo_valido = anterior  # melhor esforço

        if not ultimo_valido:
            return None
        try:
            return int(float(ultimo_valido.replace(".", "").replace(",", ".")))
        except Exception:
            return None

    def _validar_estoque(self, idx_item: int, qtd_planejada: int, nome: str) -> tuple[int, str]:
        el = self._achar_campo_estoque(idx_item)
        if el is None:
            # Não localizou o campo — segue com a quantidade planejada sem registrar obs
            return qtd_planejada, ""

        estoque = self._ler_estoque_estavel(el, timeout=5.0)
        if estoque is None:
            # Não conseguiu ler de forma confiável — não loga falsamente
            self.log(f"   ℹ️ Estoque indisponível para '{nome}' — assumindo planejado.")
            return qtd_planejada, ""

        if estoque >= qtd_planejada:
            return qtd_planejada, ""
        if estoque > 0:
            obs = f"ESTOQUE INSUFICIENTE: planejado {qtd_planejada}, disponível {estoque}, enviado {estoque}"
            self.log(f"   ⚠️ {obs} — '{nome}'")
            return estoque, obs

        obs = f"SEM ESTOQUE: planejado {qtd_planejada}, disponível 0"
        self.log(f"   🔴 {obs} — '{nome}'")
        return qtd_planejada, obs

    # ---------- salvar pedido ----------
    def _salvar_pedido(self, idx_pedido: int, max_tentativas: int = 5) -> str | None:
        for tentativa in range(1, max_tentativas + 1):
            numero = self._capturar_numero_pedido()

            try:
                self._click_js(self.wait.until(EC.element_to_be_clickable((By.ID, "botaoSalvar"))))
            except Exception as e:
                self.log(f"⚠️ Botão salvar não encontrado: {e}")
                return None

            try:
                dialog = WebDriverWait(self.driver, 8).until(
                    EC.visibility_of_element_located((By.XPATH, "//div[contains(@class,'ui-dialog')]"))
                )
                texto = _strip_accents(dialog.text).lower()
            except Exception:
                numero = self._capturar_numero_pedido() or numero
                self.log(f"✅ Pedido {idx_pedido} salvo! (Nº {numero})")
                return numero

            if any(kw in texto for kw in _ERRO_NUMERO_KEYWORDS):
                self.log(f"⚠️ Conflito nº {idx_pedido} (tentativa {tentativa}/{max_tentativas})")
                self._clicar_botao_ok(timeout=3)
                time.sleep(0.5)

                num_atual = self._capturar_numero_pedido()
                if num_atual and num_atual.isdigit():
                    novo = str(int(num_atual) + 1)
                elif numero and numero.isdigit():
                    novo = str(int(numero) + 1)
                else:
                    self.log("⚠️ Não foi possível incrementar o número.")
                    return None

                try:
                    campo = self.driver.find_element(By.ID, "numero")
                    campo.clear()
                    campo.send_keys(novo)
                    numero = novo
                    self.log(f"   Número ajustado: {num_atual} → {novo}")
                except Exception:
                    self.log("⚠️ Campo 'numero' não encontrado.")
                    return None
            else:
                self._clicar_botao_ok(timeout=3)
                time.sleep(1)
                numero = self._capturar_numero_pedido() or numero
                self.log(f"✅ Pedido {idx_pedido} salvo! (Nº {numero})")
                return numero

        self.log(f"❌ Pedido {idx_pedido}: conflito após {max_tentativas} tentativas.")
        return None

    # ---------- execução principal ----------
    def executar(self):
        """Método principal — executa todo o fluxo do bot."""
        try:
            self.carregar_planilha()

            # Inicia Chrome
            self.log("🌐 Iniciando navegador...")
            options = webdriver.ChromeOptions()
            options.add_argument("--start-maximized")
            options.add_experimental_option("detach", True)
            self.driver = webdriver.Chrome(options=options)
            self.wait = WebDriverWait(self.driver, TIMEOUT)
            self.wait_short = WebDriverWait(self.driver, 5)

            self.fazer_login()

            # Processa pedidos
            for idx_pedido, produtos in enumerate(self.pedidos, start=1):
                if self._cancelado:
                    self.log("🛑 Execução cancelada pelo usuário.")
                    break

                self.log(f"\n📋 Processando Pedido {idx_pedido}/{len(self.pedidos)} ({len(produtos)} produto(s))...")
                self.driver.get(URL_PEDIDOS)

                # Aguarda formulário limpo
                try:
                    campo_cliente = self.wait.until(EC.element_to_be_clickable((By.ID, "contato")))
                    if (campo_cliente.get_attribute("value") or "").strip():
                        self.log("⏳ Aguardando reset do formulário...")
                        self.wait.until(
                            lambda d: not (d.find_element(By.ID, "contato").get_attribute("value") or "").strip()
                        )
                        campo_cliente = self.driver.find_element(By.ID, "contato")
                except Exception as e:
                    self.log(f"⚠️ Página não carregou: {e}")
                    self.err_pedidos.append(idx_pedido)
                    continue

                # Cliente
                self.log("👤 Preenchendo cliente...")
                try:
                    campo_cliente.click()
                    campo_cliente.clear()
                    campo_cliente.send_keys(self.cliente)
                    try:
                        self.wait_short.until(EC.visibility_of_element_located((By.XPATH, AUTOCOMPLETE_XPATH)))
                        for it in self.driver.find_elements(By.XPATH, AUTOCOMPLETE_XPATH):
                            if it.is_displayed():
                                self._click_js(it)
                                break
                    except Exception:
                        campo_cliente.send_keys(Keys.ENTER)
                    self.log("✅ Cliente selecionado.")
                except Exception as e:
                    self.log(f"⚠️ Erro cliente: {e}")

                # Loja
                self.log("🏬 Selecionando loja...")
                try:
                    Select(self.wait.until(EC.element_to_be_clickable((By.ID, "loja")))).select_by_visible_text(self.loja)
                    self.log("✅ Loja selecionada.")
                except Exception as e:
                    self.log(f"⚠️ Erro loja: {e}")

                # Frete
                self.log("📦 Selecionando frete...")
                try:
                    Select(self.wait.until(EC.element_to_be_clickable((By.ID, "fretePorConta")))).select_by_visible_text(self.frete)
                    self.log("✅ Frete selecionado.")
                except Exception as e:
                    self.log(f"⚠️ Erro frete: {e}")

                # Produtos
                self.log("🧾 Inserindo produtos...")
                for i, produto in enumerate(produtos):
                    nome = produto["SKU PAI"]
                    qtd = produto["QNT PLANEJADA"]

                    try:
                        campo = self.wait.until(EC.element_to_be_clickable((By.ID, "produto_descricao")))
                        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
                        self._click_js(campo)

                        self._selecionar_sku(campo, nome)
                        time.sleep(4)  # aguarda Bling renderizar o produto e mover foco para a qty

                        qtd_enviar, obs = self._validar_estoque(i, qtd, nome)

                        qty_field = self.driver.switch_to.active_element
                        qty_field.send_keys(Keys.CONTROL, "a")
                        qty_field.send_keys(str(qtd_enviar))
                        qty_field.send_keys(Keys.TAB)
                        time.sleep(3)  # aguarda Bling persistir a quantidade

                        if obs:
                            self._alteracoes.append({"linha": produto["excel_row"], "qtd_real": qtd_enviar, "obs": obs})
                            self.log(f"⚠️ '{nome}' x{qtd_enviar} (planejado {qtd}) — {obs}")
                        else:
                            self.log(f"✅ '{nome}' x{qtd_enviar} adicionado.")

                        if i < len(produtos) - 1:
                            try:
                                btn = self.wait.until(EC.presence_of_element_located((By.ID, "add_new_item")))
                                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                                time.sleep(3)  # margem antes de clicar em "nova linha"
                                self.driver.execute_script("arguments[0].click();", btn)
                                self.log("➕ Nova linha adicionada.")
                                time.sleep(3)  # aguarda nova linha renderizar
                            except Exception as e:
                                self.log(f"⚠️ Erro nova linha: {e}")

                    except Exception as e:
                        limpo = "".join(c for c in nome if c.isalnum())[:30]
                        arq = os.path.join(self.prints_dir, f"erro_{limpo}_pedido_{idx_pedido}.png")
                        self.driver.save_screenshot(arq)
                        self.log(f"⚠️ Erro produto '{nome}': {e}")

                # Salvar
                self.log(f"💾 Salvando Pedido {idx_pedido}...")
                numero = self._salvar_pedido(idx_pedido)

                if numero is not None:
                    self.ok_pedidos.append(idx_pedido)
                    for p in produtos:
                        exist = next((a for a in self._alteracoes if a["linha"] == p["excel_row"]), None)
                        if exist:
                            exist["numero_pedido"] = numero
                        else:
                            self._alteracoes.append({"linha": p["excel_row"], "numero_pedido": numero})
                    self.log(f"   📝 Nº {numero} registrado.")
                else:
                    self.log(f"❌ Pedido {idx_pedido}: falha ao capturar nº.")
                    self.err_pedidos.append(idx_pedido)

            # Grava resultado
            self.log("\n📊 Gravando planilha de resultado...")
            try:
                self.gravar_resultado_final()
            except Exception as e:
                self.log(f"❌ Erro ao gravar resultado: {e}")

            self.log(f"\n✅ Concluído: {len(self.ok_pedidos)} salvo(s), {len(self.err_pedidos)} com erro.")
            if self.err_pedidos:
                self.log(f"❌ Pedidos com erro: {self.err_pedidos}")
            self.log("DONE")

        except Exception as e:
            if self._cancelado:
                self.log("🛑 Execução interrompida pelo usuário.")
            else:
                self.log(f"❌ Erro fatal: {e}")
            self.log("DONE")
        finally:
            drv = self.driver
            if drv is not None:
                try:
                    drv.quit()
                except Exception:
                    pass
                self.driver = None
